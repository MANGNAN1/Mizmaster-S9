from openpyxl import load_workbook
import os

# 현재 스크립트 파일의 디렉토리 경로를 가져옴
script_dir = os.path.dirname(__file__)

# 엑셀 파일 경로 설정 (현재 스크립트 파일 기준)
excel_file = os.path.join(script_dir, 'henchies.xlsx')

# 엑셀 파일 로드
wb = load_workbook(excel_file)

# 첫 번째 시트 선택
sheet = wb.active

# 변환된 데이터를 저장할 리스트
converted_data = []

# 각 행을 순회하면서 데이터를 쉼표로 구분하여 문자열로 변환
for row in sheet.iter_rows(values_only=True):
    # 각 열의 데이터를 쉼표로 연결하여 문자열로 만듦
    row_data = ','.join(str(cell) for cell in row)
    # 변환된 데이터를 리스트에 추가
    converted_data.append(row_data)

# 저장할 파일 경로 및 이름 설정
output_file = os.path.join(script_dir, 'output.txt')

# 변환된 데이터를 파일에 쓰기
with open(output_file, 'w', encoding='utf-8') as file:
    for data in converted_data:
        file.write(data + '\n')

print(f"데이터가 {output_file}에 성공적으로 저장되었습니다.")
